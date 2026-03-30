import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

EXCEL_PATH = "Attendance_Sheet_2022-2023_Masked.xlsx"

ATTENDANCE_MAP = {
    "P":    "Present",
    "WFH":  "Work From Home",
    "HWFH": "Work From Home",
    "SL":   "Sick Leave",
    "HSL":  "Sick Leave",
    "PL":   "Paid Leave",
    "HPL":  "Paid Leave",
    "FFL":  "Festival Leave",
    "HFFL": "Festival Leave",
    "BL":   "Birthday Leave",
    "BL ":  "Birthday Leave",
    "LWP":  "Leave Without Pay",
    "HLWP": "Leave Without Pay",
    "BRL":  "Bereavement Leave",
    "BRL ": "Bereavement Leave",
    "HBRL": "Bereavement Leave",
    "HBRL ":"Bereavement Leave",
    "ML":   "Menstrual Leave",
    "HML":  "Menstrual Leave",
    "WO":   "Weekly Off",
    "HO":   "Holiday",
}

PRESENCE_CODES  = {"P", "WFH", "HWFH"}      # counts toward attendance %
WFH_CODES       = {"WFH", "HWFH"}
LEAVE_CODES     = {"SL","HSL","PL","HPL","FFL","HFFL","BL","BL ","LWP","HLWP","BRL","BRL ","HBRL","HBRL ","ML","HML"}
OFF_CODES       = {"WO", "HO"}

def _safe_excel_path():
    """Try multiple locations for the Excel file."""
    candidates = [
        EXCEL_PATH,
        "/mnt/user-data/uploads/Attendance_Sheet_2022-2023_Masked.xlsx",
        os.path.join(os.path.dirname(__file__), EXCEL_PATH),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def load_attendance_data() -> pd.DataFrame:
    """
    Parse all month sheets and return a tidy long-format DataFrame:
    Columns: employee_code, name, date, code, category, month
    """
    path = _safe_excel_path()
    if path is None:
        # Return empty sample data so app still renders
        return pd.DataFrame(columns=["employee_code","name","date","code","category","month"])

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = [s for s in wb.sheetnames if s.strip() != "Attendance Key"]

    records = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        # Row 0 = dates (col 2 onwards); Row 1 = headers; Row 2+ = employees
        date_row = rows[0]
        dates = []
        for cell in date_row[2:]:
            if isinstance(cell, pd.Timestamp):
                dates.append(cell)
            elif hasattr(cell, 'date') and not isinstance(cell, str):
                try:
                    from datetime import datetime
                    dates.append(cell)
                except:
                    dates.append(None)
            elif isinstance(cell, str) and cell.startswith("="):
                dates.append(None)
            else:
                dates.append(None)

        # Determine last date column (before summary columns)
        # Summary cols start after the 31st date column
        max_date_cols = min(31, len(dates))

        for row in rows[2:]:
            if not row[0] or not isinstance(row[0], str):
                continue
            emp_code = str(row[0]).strip()
            emp_name = str(row[1]).strip() if row[1] else "Unknown"
            if emp_code.lower() in ("employee code", "employee code "):
                continue

            attendance_cells = row[2:2 + max_date_cols]
            for i, code in enumerate(attendance_cells):
                if i >= len(dates):
                    break
                date = dates[i]
                if code is None:
                    continue
                code_str = str(code).strip()
                if not code_str or code_str.startswith("="):
                    continue

                category = ATTENDANCE_MAP.get(code_str, "Other")
                records.append({
                    "employee_code": emp_code,
                    "name": emp_name,
                    "date": date,
                    "code": code_str,
                    "category": category,
                    "month": sheet_name.strip(),
                })

    df = pd.DataFrame(records)
    if df.empty:
        return df

    # Parse dates properly
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"])
    df["month_year"] = df["date"].dt.to_period("M").astype(str)
    df["day_of_week"] = df["date"].dt.day_name()
    df["is_present"]  = df["code"].isin(PRESENCE_CODES)
    df["is_wfh"]      = df["code"].isin(WFH_CODES)
    df["is_leave"]    = df["code"].isin(LEAVE_CODES)
    df["is_off"]      = df["code"].isin(OFF_CODES)

    return df


def get_employee_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate per-employee statistics across all months."""
    if df.empty:
        return pd.DataFrame()

    working_days = df[~df["is_off"]]
    
    agg = (
        working_days.groupby(["employee_code", "name"])
        .agg(
            total_days    = ("date", "count"),
            present_days  = ("is_present", "sum"),
            wfh_days      = ("is_wfh", "sum"),
            leave_days    = ("is_leave", "sum"),
        )
        .reset_index()
    )
    agg["attendance_pct"] = (agg["present_days"] / agg["total_days"] * 100).round(1)
    agg["wfh_pct"]        = (agg["wfh_days"]    / agg["total_days"] * 100).round(1)
    return agg


def get_monthly_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate per-month overview."""
    if df.empty:
        return pd.DataFrame()
    working = df[~df["is_off"]]
    agg = (
        working.groupby("month_year")
        .agg(
            total_records = ("date", "count"),
            present       = ("is_present", "sum"),
            wfh           = ("is_wfh", "sum"),
            leave         = ("is_leave", "sum"),
        )
        .reset_index()
    )
    agg["attendance_pct"] = (agg["present"] / agg["total_records"] * 100).round(1)
    agg["wfh_pct"]        = (agg["wfh"]    / agg["total_records"] * 100).round(1)
    return agg.sort_values("month_year")


def get_daily_trend(df: pd.DataFrame, month: str = None) -> pd.DataFrame:
    """Daily attendance/WFH counts, optionally filtered by month sheet name."""
    if df.empty:
        return pd.DataFrame()
    fdf = df if month is None else df[df["month"] == month]
    working = fdf[~fdf["is_off"]]
    trend = (
        working.groupby("date")
        .agg(
            present = ("is_present", "sum"),
            wfh     = ("is_wfh", "sum"),
            leave   = ("is_leave", "sum"),
            total   = ("date", "count"),
        )
        .reset_index()
        .sort_values("date")
    )
    trend["attendance_pct"] = (trend["present"] / trend["total"] * 100).round(1)
    return trend
